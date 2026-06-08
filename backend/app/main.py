"""
智能审计系统 - 主应用入口
Intelligent Audit System - Main Application Entry
"""

from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from fastapi.middleware.trustedhost import TrustedHostMiddleware
from loguru import logger
import sys
import os

from app.core.config import settings
from app.core.logging import setup_logging
from app.core.database import SessionLocal, init_db
from app.models.template import WorksheetTemplate, TemplateCategory
from app.api.v1.api import api_router

# 设置日志
setup_logging()


def _write_xlsx_template(preset: dict, file_path: str):
    """生成 xlsx 格式模板文件"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook()
    # 第一个 sheet: 审计基本信息
    ws_info = wb.active
    ws_info.title = "审计基本信息"
    info_lines = [preset["name"]]
    for line in preset["content_lines"]:
        if line.strip():
            info_lines.append(line)
    title_font = Font(name='微软雅黑', bold=True, size=14)
    for i, line in enumerate(info_lines):
        cell = ws_info.cell(row=i + 1, column=1, value=line)
        if i == 0:
            cell.font = title_font

    # 后续 sheets 来自 xlsx_sheets
    header_fill = PatternFill(start_color="E34D59", end_color="E34D59", fill_type="solid")
    header_font = Font(name='微软雅黑', bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    for sheet_def in preset.get("xlsx_sheets", []):
        ws = wb.create_sheet(title=sheet_def["name"])
        for col_idx, header in enumerate(sheet_def["headers"], 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        for row_idx, row_data in enumerate(sheet_def["rows"], 2):
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border

        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len * 3 + 4, 50)

    wb.save(file_path)


def init_preset_templates():
    """初始化预设底稿模板（生成 xlsx 格式，下载时可转为 CSV）"""
    import os

    db = SessionLocal()
    try:
        template_dir = os.path.join(settings.UPLOAD_DIR, "templates")
        os.makedirs(template_dir, exist_ok=True)

        # 先删除旧预设模板记录和文件，确保新格式生效
        old_presets = db.query(WorksheetTemplate).filter(WorksheetTemplate.is_preset == True).all()
        for old in old_presets:
            if os.path.exists(old.file_path):
                os.remove(old.file_path)
                logger.info(f"已删除旧模板文件: {old.file_path}")
            db.delete(old)
        db.flush()

        # ==================== 4 个 xlsx 模板 ====================
        xlsx_presets = [
            {
                "id": "TMP-P01",
                "name": "费用审计底稿模板",
                "category": TemplateCategory.FINANCIAL.value,
                "file_type": "xlsx",
                "description": "适用于差旅费、招待费、办公费等费用科目的标准化审计底稿，含费用明细表、审批链路核查、异常标识等标准章节",
                "content_lines": [
                    "一、审计基本信息",
                    "被审计单位：______________",
                    "审计期间：______________",
                    "审计人员：______________",
                    "审计日期：______________",
                    "",
                    "二、审计目的",
                    "对费用支出的真实性、合规性、合理性进行审查。",
                    "",
                    "三、费用明细核查",
                    "详见表单数据，逐笔核查凭证号、金额、报销人、审批链路。",
                    "",
                    "四、审计结论",
                    "□ 合规  □ 基本合规  □ 存在缺陷  □ 重大缺陷",
                ],
                "xlsx_sheets": [
                    {
                        "name": "费用明细表",
                        "headers": ["序号", "凭证号", "费用类型", "金额(元)", "报销人", "审批人", "是否合规", "备注"],
                        "rows": [["1", "", "", "", "", "", "", ""]],
                    },
                    {
                        "name": "异常发现",
                        "headers": ["异常类型", "具体描述", "涉及金额(元)", "整改建议"],
                        "rows": [
                            ["超标准报销", "", "", ""],
                            ["缺少附件", "", "", ""],
                            ["重复报销", "", "", ""],
                            ["其他异常", "", "", ""],
                        ],
                    },
                ],
            },
            {
                "id": "TMP-P02",
                "name": "采购审计底稿模板",
                "category": TemplateCategory.PURCHASE.value,
                "file_type": "xlsx",
                "description": "适用于采购流程合规性审计的标准化底稿，含供应商资质审查、招标流程合规性、合同条款审查等标准章节",
                "content_lines": [
                    "一、审计基本信息",
                    "被审计单位：______________",
                    "审计期间：______________",
                    "采购类型：□ 工程类  □ 物资类  □ 服务类",
                    "审计人员：______________",
                    "",
                    "二、审计目的",
                    "对采购流程的合规性、供应商资质的真实性进行审查。",
                    "",
                    "三、审计要点",
                    "1. 供应商资质是否齐全有效",
                    "2. 招标流程是否合规",
                    "3. 合同条款是否完整合理",
                    "",
                    "四、审计结论",
                    "□ 合规  □ 基本合规  □ 存在缺陷  □ 重大缺陷",
                ],
                "xlsx_sheets": [
                    {
                        "name": "供应商资质审查",
                        "headers": ["序号", "供应商名称", "营业执照", "资质证书", "业绩证明", "信用记录", "审查结论"],
                        "rows": [["1", "", "", "", "", "", ""]],
                    },
                    {
                        "name": "招标流程核查",
                        "headers": ["核查项", "核查内容", "是否符合", "备注"],
                        "rows": [
                            ["招标方式", "□公开招标 □邀请招标 □竞争性谈判 □单一来源", "", ""],
                            ["招标文件完整性", "", "", ""],
                            ["评标过程规范性", "", "", ""],
                        ],
                    },
                    {
                        "name": "合同条款审查",
                        "headers": ["审查项", "审查内容", "是否存在风险", "备注"],
                        "rows": [
                            ["合同金额是否超预算", "", "", ""],
                            ["付款条款是否合理", "", "", ""],
                            ["违约责任条款是否明确", "", "", ""],
                        ],
                    },
                ],
            },
            {
                "id": "TMP-P03",
                "name": "资金审计底稿模板",
                "category": TemplateCategory.FUND.value,
                "file_type": "xlsx",
                "description": "适用于货币资金、银行余额调节、资金划转等审计的标准化底稿，含银行对账、大额资金流向追踪、权限审批等标准章节",
                "content_lines": [
                    "一、审计基本信息",
                    "被审计单位：______________",
                    "审计期间：______________",
                    "审计人员：______________",
                    "",
                    "二、审计目的",
                    "对货币资金的真实性、完整性、安全性进行审查。",
                    "",
                    "三、审计要点",
                    "1. 银行账户余额核对",
                    "2. 大额资金流向追踪（单笔≥10万元）",
                    "3. 资金划转审批权限核验",
                    "4. 关联方资金往来排查",
                    "",
                    "四、审计结论",
                    "□ 合规  □ 基本合规  □ 存在缺陷  □ 重大缺陷",
                ],
                "xlsx_sheets": [
                    {
                        "name": "银行账户清单",
                        "headers": ["序号", "开户行", "账号", "币种", "账面余额", "银行对账单余额", "差异"],
                        "rows": [["1", "", "", "", "", "", ""]],
                    },
                    {
                        "name": "大额资金流向追踪",
                        "headers": ["日期", "凭证号", "交易对方", "金额(元)", "资金用途", "审批人", "是否合规"],
                        "rows": [["", "", "", "", "", "", ""]],
                    },
                    {
                        "name": "银行余额调节表",
                        "headers": ["项目", "金额(元)", "说明"],
                        "rows": [
                            ["企业账面余额", "", ""],
                            ["加：银行已收企业未收", "", ""],
                            ["减：银行已付企业未付", "", ""],
                            ["调整后余额", "", ""],
                        ],
                    },
                ],
            },
            {
                "id": "TMP-P04",
                "name": "资产盘点审计底稿模板",
                "category": TemplateCategory.ASSET.value,
                "file_type": "xlsx",
                "description": "适用于固定资产、存货盘点审计的标准化底稿，含资产清单核对、盘点差异分析、折旧计提核查等标准章节",
                "content_lines": [
                    "一、审计基本信息",
                    "被审计单位：______________",
                    "盘点日期：______________",
                    "资产类别：□ 固定资产  □ 存货  □ 无形资产",
                    "审计人员：______________",
                    "",
                    "二、审计目的",
                    "对资产的存在性、完整性、计价准确性进行审查。",
                    "",
                    "三、审计要点",
                    "1. 资产清单与实际盘点核对",
                    "2. 盘盈盘亏差异分析",
                    "3. 折旧计提方法及年限核查",
                    "",
                    "四、审计结论",
                    "□ 合规  □ 基本合规  □ 存在缺陷  □ 重大缺陷",
                ],
                "xlsx_sheets": [
                    {
                        "name": "资产清单核对",
                        "headers": ["序号", "资产编号", "资产名称", "账面数量", "实盘数量", "差异", "差异原因"],
                        "rows": [["1", "", "", "", "", "", ""]],
                    },
                    {
                        "name": "盘点差异分析",
                        "headers": ["差异类型", "资产名称", "差异数量", "差异金额(元)", "原因分析"],
                        "rows": [
                            ["盘盈", "", "", "", ""],
                            ["盘亏", "", "", "", ""],
                            ["闲置", "", "", "", ""],
                        ],
                    },
                    {
                        "name": "折旧计提核查",
                        "headers": ["核查项", "核查内容", "是否符合", "备注"],
                        "rows": [
                            ["折旧方法是否一致", "", "", ""],
                            ["折旧年限是否合理", "", "", ""],
                            ["残值率是否符合规定", "", "", ""],
                        ],
                    },
                ],
            },
        ]

        for preset in xlsx_presets:
            file_type = preset["file_type"]
            file_name = f"{preset['name']}.{file_type}"
            file_path = os.path.join(template_dir, f"{preset['id']}_{file_name}")

            _write_xlsx_template(preset, file_path)
            file_size = os.path.getsize(file_path)

            template = WorksheetTemplate(
                id=preset["id"],
                name=preset["name"],
                category=preset["category"],
                description=preset["description"],
                file_name=file_name,
                file_path=file_path,
                file_size=file_size,
                file_type=file_type,
                is_preset=True,
                download_count=0,
                created_by="system",
            )
            db.add(template)
            logger.info(f"预设模板已初始化: {preset['id']} - {preset['name']} (.{file_type})")

        db.commit()
        logger.info(f"预设模板初始化完成，共 {len(xlsx_presets)} 个（xlsx 格式，下载时可转为 CSV）")
    except Exception as e:
        db.rollback()
        logger.warning(f"预设模板初始化失败: {e}")
    finally:
        db.close()


# 创建FastAPI应用
app = FastAPI(
    title="智能审计系统 API",
    description="Intelligent Audit System - 数据驱动、智能作业",
    version="1.0.0",
    openapi_url=f"{settings.API_V1_STR}/openapi.json",
    docs_url="/docs",
    redoc_url="/redoc",
)

# CORS中间件配置
if settings.BACKEND_CORS_ORIGINS:
    app.add_middleware(
        CORSMiddleware,
        allow_origins=[str(origin) for origin in settings.BACKEND_CORS_ORIGINS],
        allow_credentials=True,
        allow_methods=["*"],
        allow_headers=["*"],
    )

# 添加自定义中间件（暂不启用，待修复）
# app.add_middleware(TrustedHostMiddleware, allowed_hosts=["*"])
# app.add_middleware(LoggingMiddleware)
# app.add_middleware(AuthMiddleware)

# 包含API路由
app.include_router(api_router, prefix=settings.API_V1_STR)


@app.on_event("startup")
async def startup_event():
    """应用启动事件"""
    logger.info("=" * 60)
    logger.info("智能审计系统启动中...")
    logger.info(f"环境: {settings.ENVIRONMENT}")
    logger.info(f"调试模式: {settings.DEBUG}")
    logger.info("=" * 60)
    
    # 初始化数据库表
    init_db()
    
    # 初始化预设模板
    init_preset_templates()


@app.on_event("shutdown")
async def shutdown_event():
    """应用关闭事件"""
    logger.info("智能审计系统关闭中...")


@app.get("/")
async def root():
    """根路径 - 健康检查"""
    return {
        "system": "智能审计系统",
        "version": "1.0.0",
        "status": "running",
        "docs": "/docs",
        "environment": settings.ENVIRONMENT,
    }


@app.get("/health")
async def health_check():
    """健康检查端点"""
    return {
        "status": "healthy",
        "version": "1.0.0",
    }


if __name__ == "__main__":
    import uvicorn
    
    uvicorn.run(
        "main:app",
        host="0.0.0.0",
        port=8000,
        reload=settings.DEBUG,
        log_level="info" if settings.DEBUG else "warning",
    )
