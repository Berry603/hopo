"""
审计程序执行结果导出为 Excel，自动归档到项目文件夹
"""
import os
from pathlib import Path
from datetime import datetime
from loguru import logger

# 项目根目录
AUDIT_PROJECTS_ROOT = Path(__file__).resolve().parent.parent.parent.parent / "Auditoprojects"


def generate_procedure_excel(
    project_code: str,
    project_name: str,
    procedure_code: str,
    procedure_name: str,
    items: list,
    rows: list,
    conclusion: str = "",
) -> str:
    """
    生成审计程序执行结果 Excel 并保存到项目文件夹
    
    Returns:
        file_path: 生成的 Excel 文件绝对路径
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    safe_name = _sanitize(project_name)
    project_dir = AUDIT_PROJECTS_ROOT / f"{project_code}_{safe_name}"
    # 穿行测试文件存到 05_测试与底稿/01_穿行测试/
    target_dir = project_dir / "05_测试与底稿" / "01_穿行测试"
    target_dir.mkdir(parents=True, exist_ok=True)
    
    filename = f"{procedure_code}_{_sanitize(procedure_name)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    file_path = str(target_dir / filename)
    
    wb = Workbook()
    
    # ===== Sheet1: 检查结果明细 =====
    ws = wb.active
    ws.title = "检查结果明细"
    
    # 样式
    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="D7011D", end_color="D7011D", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="微软雅黑", size=10)
    cell_align = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    
    # 程序信息行
    ws.merge_cells("A1:H1")
    ws["A1"] = f"程序名称：{procedure_code} - {procedure_name}"
    ws["A1"].font = Font(name="微软雅黑", bold=True, size=14, color="D7011D")
    
    ws.merge_cells("A2:H2")
    ws["A2"] = f"生成日期：{datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(name="微软雅黑", size=9, color="666666")
    
    # 表头
    field_labels = [item.field_label for item in (items or [])]
    headers = ["序号"] + field_labels + ["结论", "备注"]
    start_row = 4
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    # 数据行
    for row_idx, row in enumerate(rows, 1):
        excel_row = start_row + row_idx
        
        # 序号
        cell = ws.cell(row=excel_row, column=1, value=row_idx)
        cell.font = cell_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
        # 字段值
        row_data = row.data or {}
        for field_idx, item in enumerate(items or []):
            val = row_data.get(item.field_name, "")
            # 处理布尔值
            if isinstance(val, bool):
                val = "是" if val else "否"
            cell = ws.cell(row=excel_row, column=field_idx + 2, value=val)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border
            
            # 异常行标红
            if row.conclusion == "异常":
                cell.fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
        
        # 结论
        conclusion_cell = ws.cell(row=excel_row, column=len(headers) - 1, value=row.conclusion or "")
        conclusion_cell.font = Font(name="微软雅黑", size=10,
                                     color="FF0000" if row.conclusion == "异常" else (
                                         "00AA00" if row.conclusion == "正常" else "666666"))
        conclusion_cell.alignment = Alignment(horizontal="center", vertical="center")
        conclusion_cell.border = thin_border
        
        # 备注
        remark_cell = ws.cell(row=excel_row, column=len(headers), value=row.remark or "")
        remark_cell.font = cell_font
        remark_cell.alignment = cell_align
        remark_cell.border = thin_border
    
    # 设置列宽
    ws.column_dimensions["A"].width = 6
    for i, item in enumerate(items or []):
        ws.column_dimensions[chr(66 + i)].width = max(15, len(item.field_label) * 2)
    ws.column_dimensions[chr(66 + len(items or []))].width = 10
    ws.column_dimensions[chr(67 + len(items or []))].width = 20
    
    # ===== Sheet2: 测试结论汇总 =====
    ws2 = wb.create_sheet("测试结论汇总")
    ws2.merge_cells("A1:B1")
    ws2["A1"] = "测试结论"
    ws2["A1"].font = Font(name="微软雅黑", bold=True, size=14, color="D7011D")
    
    summary_data = [
        ("程序编号", procedure_code),
        ("程序名称", procedure_name),
        ("样本数量", str(len(rows))),
        ("正常", str(sum(1 for r in rows if r.conclusion == "正常"))),
        ("异常", str(sum(1 for r in rows if r.conclusion == "异常"))),
        ("待确认", str(sum(1 for r in rows if r.conclusion == "待确认"))),
        ("测试结论", conclusion),
        ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for i, (label, value) in enumerate(summary_data, 3):
        ws2.cell(row=i, column=1, value=label).font = Font(name="微软雅黑", bold=True, size=10)
        ws2.cell(row=i, column=2, value=value).font = Font(name="微软雅黑", size=10)
    
    ws2.column_dimensions["A"].width = 15
    ws2.column_dimensions["B"].width = 50
    
    wb.save(file_path)
    logger.info(f"程序执行Excel已生成: {file_path}")
    return file_path


def _sanitize(name: str) -> str:
    """清理文件名中的非法字符"""
    illegal = r'<>:"/\|?*'
    for c in illegal:
        name = name.replace(c, "_")
    return name.strip().replace(" ", "_")[:50]
