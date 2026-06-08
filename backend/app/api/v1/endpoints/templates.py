"""
底稿模板管理 API
Worksheet Template Management API
"""

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from loguru import logger
import os
import uuid
import shutil
import io
import csv
import urllib.parse
from typing import Optional, List
from datetime import datetime

from app.core.database import get_db
from app.core.config import settings
from app.models.template import WorksheetTemplate, TemplateCategory, CATEGORY_LABELS

router = APIRouter()

# 确保上传目录存在
TEMPLATE_DIR = os.path.join(settings.UPLOAD_DIR, "templates")
os.makedirs(TEMPLATE_DIR, exist_ok=True)


def get_next_template_id(db: Session) -> str:
    """生成下一个模板ID"""
    last = db.query(WorksheetTemplate).order_by(WorksheetTemplate.id.desc()).first()
    if not last:
        return "TMP-001"
    try:
        num = int(last.id.split("-")[1]) + 1
        return f"TMP-{num:03d}"
    except (IndexError, ValueError):
        return f"TMP-{uuid.uuid4().hex[:6].upper()}"


@router.get("/templates")
def list_templates(
    category: Optional[str] = Query(None, description="按分类筛选"),
    keyword: Optional[str] = Query(None, description="按名称搜索"),
    file_type: Optional[str] = Query(None, description="按文件类型筛选"),
    db: Session = Depends(get_db),
):
    """获取底稿模板列表"""
    query = db.query(WorksheetTemplate)
    
    if category and category != "all":
        query = query.filter(WorksheetTemplate.category == category)
    
    if keyword:
        query = query.filter(WorksheetTemplate.name.ilike(f"%{keyword}%"))
    
    if file_type:
        query = query.filter(WorksheetTemplate.file_type == file_type)
    
    templates = query.order_by(WorksheetTemplate.is_preset.desc(), WorksheetTemplate.created_at.desc()).all()
    
    return {
        "total": len(templates),
        "items": [t.to_dict() for t in templates],
    }


@router.get("/templates/categories")
def list_categories():
    """获取模板分类列表"""
    return {
        "items": [
            {"value": cat.value, "label": CATEGORY_LABELS[cat]}
            for cat in TemplateCategory
        ]
    }


@router.get("/templates/{template_id}")
def get_template(template_id: str, db: Session = Depends(get_db)):
    """获取模板详情"""
    template = db.query(WorksheetTemplate).filter(WorksheetTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail=f"模板 {template_id} 不存在")
    return template.to_dict()


@router.post("/templates/upload")
async def upload_template(
    name: str = Form(..., description="模板名称"),
    category: str = Form(default=TemplateCategory.OTHER.value, description="模板分类"),
    description: Optional[str] = Form(default=None, description="模板描述"),
    file: UploadFile = File(..., description="模板文件"),
    db: Session = Depends(get_db),
):
    """上传底稿模板文件"""
    # 验证文件扩展名
    if not file.filename:
        raise HTTPException(status_code=400, detail="文件名不能为空")
    
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in settings.ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"不支持的文件格式: {ext}。支持的格式: {', '.join(settings.ALLOWED_EXTENSIONS)}"
        )
    
    # 检查文件大小
    content = await file.read()
    file_size = len(content)
    if file_size > settings.MAX_UPLOAD_SIZE:
        raise HTTPException(
            status_code=400,
            detail=f"文件大小 {file_size / 1024 / 1024:.1f}MB 超过限制 {settings.MAX_UPLOAD_SIZE / 1024 / 1024:.0f}MB"
        )
    
    # 验证分类
    valid_categories = [c.value for c in TemplateCategory]
    if category not in valid_categories:
        category = TemplateCategory.OTHER.value
    
    # 生成唯一文件名并保存
    template_id = get_next_template_id(db)
    stored_name = f"{template_id}_{uuid.uuid4().hex[:8]}{ext}"
    file_path = os.path.join(TEMPLATE_DIR, stored_name)
    
    with open(file_path, "wb") as f:
        f.write(content)
    
    # 创建数据库记录
    template = WorksheetTemplate(
        id=template_id,
        name=name,
        category=category,
        description=description,
        file_name=file.filename,
        file_path=file_path,
        file_size=file_size,
        file_type=ext.lstrip("."),
        is_preset=False,
        download_count=0,
        created_by="admin",
    )
    
    db.add(template)
    db.commit()
    db.refresh(template)
    
    logger.info(f"模板上传成功: {template_id} - {name} ({file.filename}, {file_size} bytes)")
    
    return {"success": True, "data": template.to_dict(), "message": f"模板「{name}」上传成功"}


@router.get("/templates/{template_id}/download")
def download_template(
    template_id: str,
    format: Optional[str] = Query(None, description="下载格式: xlsx 或 csv"),
    db: Session = Depends(get_db),
):
    """下载模板文件（支持 xlsx/csv 格式切换）"""
    template = db.query(WorksheetTemplate).filter(WorksheetTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail=f"模板 {template_id} 不存在")

    # 确定目标格式：参数优先，否则用原始格式
    target_format = format or template.file_type

    # 更新下载次数
    template.download_count = (template.download_count or 0) + 1
    db.commit()

    # --- CSV 下载（从 xlsx 按需生成）---
    if target_format == "csv":
        if not os.path.exists(template.file_path):
            raise HTTPException(status_code=404, detail="模板文件不存在，可能已被删除")

        try:
            from openpyxl import load_workbook

            # 直接生成 CSV 字节流
            output = io.BytesIO()
            output.write(b'\xef\xbb\xbf')  # UTF-8 BOM
            wrapper = io.TextIOWrapper(output, encoding='utf-8', newline='')

            writer = csv.writer(wrapper)
            wb = load_workbook(template.file_path, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                writer.writerow([f"--- {sheet_name} ---"])
                for row in ws.iter_rows(values_only=True):
                    writer.writerow([str(cell) if cell is not None else "" for cell in row])
                writer.writerow([])
            wb.close()
            wrapper.flush()
            wrapper.detach()

            csv_bytes = output.getvalue()
            output.close()

            csv_filename = template.file_name.rsplit(".", 1)[0] + ".csv"
            encoded_filename = urllib.parse.quote(csv_filename)
            return StreamingResponse(
                io.BytesIO(csv_bytes),
                media_type="text/csv; charset=utf-8",
                headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"},
            )
        except Exception as e:
            logger.error(f"生成 CSV 失败: {e}")
            raise HTTPException(status_code=500, detail=f"CSV 生成失败: {str(e)}")

    # --- xlsx 下载（原始文件）---
    if not os.path.exists(template.file_path):
        raise HTTPException(status_code=404, detail="模板文件不存在，可能已被删除")

    media_type_map = {
        "pdf": "application/pdf",
        "doc": "application/msword",
        "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "xls": "application/vnd.ms-excel",
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "png": "image/png",
        "jpg": "image/jpeg",
        "jpeg": "image/jpeg",
        "csv": "text/csv",
    }
    media_type = media_type_map.get(template.file_type, "application/octet-stream")

    return FileResponse(
        path=template.file_path,
        filename=template.file_name,
        media_type=media_type,
    )


@router.delete("/templates/{template_id}")
def delete_template(template_id: str, db: Session = Depends(get_db)):
    """删除模板"""
    template = db.query(WorksheetTemplate).filter(WorksheetTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail=f"模板 {template_id} 不存在")
    
    if template.is_preset:
        raise HTTPException(status_code=403, detail="系统预设模板不可删除")
    
    # 删除文件
    if os.path.exists(template.file_path):
        os.remove(template.file_path)
    
    # 删除记录
    db.delete(template)
    db.commit()
    
    logger.info(f"模板已删除: {template_id} - {template.name}")
    
    return {"success": True, "message": f"模板「{template.name}」已删除"}
